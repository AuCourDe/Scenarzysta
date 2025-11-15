"""
Moduł do eksportu scenariuszy testowych do formatu Excel.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from pathlib import Path

class ExcelExporter:
    """
    Klasa do eksportu scenariuszy testowych do Excel.
    """
    
    def __init__(self):
        """Inicjalizacja eksportera."""
        pass
    
    def export(self, test_cases: List[Dict], output_path: str, project_name: str = "Projekt") -> str:
        """
        Eksportuje scenariusze testowe do pliku Excel.
        
        Args:
            test_cases: Lista słowników z przypadkami testowymi
            output_path: Ścieżka do pliku wyjściowego
            project_name: Nazwa projektu (dla nagłówka)
            
        Returns:
            Ścieżka do wygenerowanego pliku
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Scenariusze Testowe"
        
        # Styl nagłówka
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Styl komórek
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        
        # Nagłówki kolumn
        headers = [
            "Test Case ID",
            "Nazwa scenariusza",
            "Krok do wykonania",
            "Wymaganie",
            "Rezultat oczekiwany"
        ]
        
        # Dodaj nagłówki
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Dodaj dane
        for row_idx, test_case in enumerate(test_cases, start=2):
            # Test Case ID
            test_id = test_case.get('test_case_id', f"TC_{row_idx-1:04d}")
            ws.cell(row=row_idx, column=1, value=test_id).border = border
            ws.cell(row=row_idx, column=1).alignment = cell_alignment
            
            # Nazwa scenariusza
            scenario_name = test_case.get('scenario_name', 'N/A')
            ws.cell(row=row_idx, column=2, value=scenario_name).border = border
            ws.cell(row=row_idx, column=2).alignment = cell_alignment
            
            # Krok do wykonania
            step_action = test_case.get('step_action', 'N/A')
            ws.cell(row=row_idx, column=3, value=step_action).border = border
            ws.cell(row=row_idx, column=3).alignment = cell_alignment
            
            # Wymaganie
            requirement = test_case.get('requirement', 'N/A')
            ws.cell(row=row_idx, column=4, value=requirement).border = border
            ws.cell(row=row_idx, column=4).alignment = cell_alignment
            
            # Rezultat oczekiwany
            expected_result = test_case.get('expected_result', 'N/A')
            ws.cell(row=row_idx, column=5, value=expected_result).border = border
            ws.cell(row=row_idx, column=5).alignment = cell_alignment
        
        # Ustaw szerokość kolumn
        ws.column_dimensions['A'].width = 20  # Test Case ID
        ws.column_dimensions['B'].width = 30  # Nazwa scenariusza
        ws.column_dimensions['C'].width = 50  # Krok do wykonania
        ws.column_dimensions['D'].width = 30  # Wymaganie
        ws.column_dimensions['E'].width = 50  # Rezultat oczekiwany
        
        # Ustaw wysokość wierszy nagłówka
        ws.row_dimensions[1].height = 30
        
        # Zapisz plik
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path)
